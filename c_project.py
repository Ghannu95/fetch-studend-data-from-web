from tkinter import *
from PIL import ImageTk, Image
from tkinter import  messagebox
from tkinter import filedialog, ttk
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import chromedriver_binary,time
import requests
import openpyxl
from openpyxl import Workbook
import mysql.connector
import pandas as pd


root = Tk()
root.title("<title of the project>")
root.iconbitmap('<address of the icon file>')
root.geometry("1520x820")
#root.resizable(width=0, height=0)



frame3 = LabelFrame(root, padx=5, width = 1000, height = 400)
frame3.grid(row=0, column=1, padx=5, pady=5)
frame3.grid_propagate(False)

#fame 4
frame4 = LabelFrame(root, padx=5, width=1000, height=400)
frame4.grid(row=1, column=1, padx=5, pady=5)
frame4.grid_propagate(False)
label2 = Label(frame4, text="Scrapped Data...", font=("Helvetica", 24))
label2.grid(row=0, column=0, pady=5, sticky=W, columnspan=3)

global collage_name
college_name = "<collage name>"
driver = None
def upload():
    global filename
    global driver
    if not driver:
        Path = "C:\Program Files (x86)\chromedriver.exe"
        driver = webdriver.Chrome(Path)

        def login():
            driver.get("https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin")
            elem = driver.find_element_by_id("username")
            elem.send_keys("<username of LinkedIn>")
            password = "<password>"
            elem2 = driver.find_elements_by_id('password')
            elem2[0].send_keys(password)
            driver.find_element_by_tag_name('button').click()
            with open("file.txt", "w") as file:
                file.write("1")

        login()
        extract_names(filename)


db = mysql.connector.connect(
        host="<localhost>",
        user="<user>",
        passwd="<password>",
        database="<name of database>"
    )
mycursor = db.cursor()

def store(dict, linkurl, fullname):
    global db

    with open("file.txt", "r") as file:
        num = int(file.read())
    with open("file.txt", "w") as file:
        file.write(str(num + 1))

    global mycursor
    mycursor = db.cursor()
    insert1 = (
            "INSERT INTO student(stuID,stuname,curr_status,linkedinlink)"// you have to use the name of table that is in your data base
            "VALUES (%s, %s, %s, %s)"
    )
    data1 = (num, dict['t1'][0], dict['t1'][1], linkurl)
    insert2 = (
            "INSERT INTO stu_education(stuname,institute)"// same you have to use your table
            "VALUES (%s, %s)"
        )
    mycursor.execute(insert1, data1)
    data2 = (fullname, "")
    if (len(dict['t2']) == 0):
        mycursor.execute(insert2, data2)
    else:
        for li in dict['t2']:
            data2 = (fullname, li[0] + " " + li[1])
            mycursor.execute(insert2, data2)
    db.commit()



def Get_data(links, fullname):
    i = 0
    for i in range(len(links)):
        driver.get(links[i])
        time.sleep(6)
        curr_status = driver.find_elements_by_css_selector('.lt-line-clamp--multi-line')[0].text
        time.sleep(1)
        inst_names = driver.find_elements_by_css_selector('#education-section h3')
        time.sleep(3)
        duration = driver.find_elements_by_css_selector('#education-section .t-black--light span')
        time.sleep(2)
        print(curr_status)
        li1 = []
        li1.append(fullname)
        li1.append(curr_status)
        dict = {}
        dict['t1'] = li1
        j = 0
        li_check = [college_name, "2014 - 2018"]
        for dur in duration:
            if (dur.text == "Dates attended or expected graduation"):
                duration.remove(dur)
        li3 = []
        flag = False
        for j in range(len(inst_names)):
            li2 = []
            li2.append(inst_names[j].text)
            li2.append(duration[j].text)
            li3.append(li2)
        dict['t2'] = li3

        #print(dict)
        # if(flag==True):
        store(dict, driver.current_url, fullname)

def extract_names(filename):
    b = list(filename)
    for i in range(0, len(b)):
        if b[i] == '/':
            b[i] = '\\\\'
    filename = ''.join(b)

    wb = openpyxl.load_workbook(filename)
    sheetList = wb.sheetnames
    sh = wb[sheetList[0]]
    for i in range(1, sh.max_row):
        fullname = sh.cell(i, 3).value[:]
        ind = fullname.find(' ')
        first_name = fullname[0:ind]
        sir_name = fullname[ind + 1:]
        driver.get(
            f"https://www.linkedin.com/search/results/people/?keywords={first_name}%20{sir_name} <you have to enter some keyword for searching of your institution>&origin=CLUSTER_EXPANSION")
        time.sleep(1)
        ans = driver.find_elements_by_css_selector('.pb3 .t-16 a')
        i = 0
        li = []
        for i in range(len(ans)):
            li.append(ans[i].get_attribute('href'))
        Get_data(li, fullname)
    on_close()

def on_close():
    lbl1 = Label(frame3, text = "Scrapping Completed, Click Show button to show data", font=("Helvetica", 16))
    lbl1.grid(row = 3, column = 0, pady = 10)
    global driver

    if driver:
        driver.close()
        driver = None



style = ttk.Style()
style.theme_use("clam")

style.configure("Treeview",
                background="white",
                foreground="black",
                rowheight=25,
                fieldbackgeound="white"
                )
style.map('Treeview',
          background=[('selected', 'blue')])




my_label = Label(frame3, text="").grid()


tree_frame1 = Frame(frame3, width = 600, height=450)
tree_frame1.grid(row = 1, column = 0, columnspan = 5, padx=10)
#tree_frame1.pack()

tree_frame2 = Frame(frame4, width =600, height=450)
tree_frame2.grid(row =1, column =0, columnspan = 5)


my_tree1 = ttk.Treeview(tree_frame1)
my_tree2 = ttk.Treeview(tree_frame2)


#file open function
def file_open():
    global filename
    filename = filedialog.askopenfilename(
        initialdir ="<address of your file manager>",
        title = "Open A File",
        filetype=(("xlsx files", "*.xlsx"), ("All Files", "*.*"))
    )

    if filename:
        try:
            filename = r"{}".format(filename)

            df1 = pd.read_excel(filename)
            file = df1
        except ValueError:
            my_label.config(text="File couldn't be Opened...try again")
        except FileNotFoundError:
            my_label.config(text="File couldn't be found...")



    clear_tree1()

    tree_scroll = Scrollbar(tree_frame1)
    tree_scroll.pack(side=RIGHT, fill=Y)
    my_tree1.configure(yscrollcommand=tree_scroll.set)
    tree_scroll.config(command=my_tree1.yview)

    tree_scroll = Scrollbar(tree_frame1, orient=HORIZONTAL)
    tree_scroll.pack(side=BOTTOM, fill=X)
    my_tree1.configure(xscrollcommand=tree_scroll.set)
    tree_scroll.config(command=my_tree1.xview)


    # setup new tree view
    my_tree1["column"] = list(df1.columns)
    my_tree1["show"] = "headings"
    #loop through columns
    for column in my_tree1["column"]:
        my_tree1.heading(column, text=column)
        my_tree1.column(column, anchor='center', width=84, minwidth=100)


    my_tree1.tag_configure('oddrow', background= "white")
    my_tree1.tag_configure('evenrow', background="lightblue")

    # put data in tree view
    global count
    count = 0
    df_rows = df1.to_numpy().tolist()
    for row in df_rows:
        if count % 2 == 0:
            my_tree1.insert("", "end", values=row, tags=('evenrow',))
        else:
            my_tree1.insert("", "end", values=row, tags=('oddrow',))
        count += 1

     # pack the treeview
    #my_tree.grid(row=1, column=0, padx=10, pady=30, sticky=(N, S, E, W), columnspan=50)
    my_tree1.pack(side = LEFT)


def clear_tree1():
    my_tree1.delete(*my_tree1.get_children())


result = list()
def show_data(exl_file):

    exl_path = "<path of file that is saved in filemanager using save file function>" + exl_file
    df2 = pd.read_excel(exl_path)

    clear_tree2()

    tree_scroll = Scrollbar(tree_frame2)
    tree_scroll.pack(side=RIGHT, fill=Y)
    my_tree1.configure(yscrollcommand=tree_scroll.set)
    tree_scroll.config(command=my_tree2.yview)

    tree_scroll = Scrollbar(tree_frame2, orient=HORIZONTAL)
    tree_scroll.pack(side=BOTTOM, fill=X)
    my_tree1.configure(xscrollcommand=tree_scroll.set)
    tree_scroll.config(command=my_tree2.xview)


   # setup new tree view
    my_tree2["column"] = list(df2.columns)
    my_tree2["show"] = "headings"
    #loop through columns
    for column in my_tree2["column"]:
        my_tree2.heading(column, text=column)
        my_tree2.column(column, anchor='center', width=84, minwidth=95)


    my_tree2.tag_configure('oddrow', background= "lightblue")
    my_tree2.tag_configure('evenrow', background="white")

   # put data in tree view
    global count
    count = 0
    global df_rows
    df_rows = df2.to_numpy().tolist()
    for row in df_rows:
        if count % 2 == 0:
            my_tree2.insert("", "end", values=row, tags=('evenrow',))
        else:
            my_tree2.insert("", "end", values=row, tags=('oddrow',))
        count += 1

    # pack the treeview
    #my_tree2.grid(row=1, column=0, padx=10, pady=30, sticky=(N, S, E, W), columnspan=50)
    my_tree2.pack(side= LEFT)

def clear_tree2():
    my_tree2.delete(*my_tree2.get_children())

#df_rows = list()

def my_delete():
    global mycursor
    mycursor.execute("TRUNCATE TABLE student")
    mycursor.execute("TRUNCATE TABLE stu_education")
    db.commit()

    win_exit()
    #show_data() # refresh the window with new records

def win_exit():
    root.quit()

def convert(csvfile):

        exl_file = entry.get() + '.xlsx'
        wb = Workbook()
        ws = wb.active
        with open(csvfile, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save('{}'.format(exl_file))

        show_data(exl_file)


# write to csv


def write_to_csv(result):
    csv_file = entry.get() + ".csv"
    mycursor.execute("SELECT * FROM student, stu_education")
    result = mycursor.fetchall()
    with open('{}'.format(csv_file), 'a', newline="") as f:
        w = csv.writer(f, dialect='excel')
        for record in result:
            w.writerow(record)


    csvfile = "<path of file that is saved in filemanager using save file function>" + csv_file

    convert(csvfile)




#frame 1

frame1 = LabelFrame(root, padx=19, pady=10, width=500, height=400)
frame1.grid(row=0, column=0)
frame1.grid_propagate(False)


label3 = Label(frame3, text="Uploaded Data...", font=("Helvetica", 24))
label3.grid(row=0, column=0, pady=5, sticky=W)



open_btn = PhotoImage(file='<address of photo image of file open button>')
img_lable1 = Label(image=open_btn)
b= Button(frame1, image=open_btn, command=file_open, borderwidth=0)
b.grid(row=2, column=1, pady=60)

upload_btn = PhotoImage(file='<address of upload button image>')
img_lable2 = Label(image=upload_btn)
upload_Button = Button(frame1, image=upload_btn, borderwidth=0, command=upload)
upload_Button.grid(row=5, column=1, pady=40)


label1 = Label(frame1, text="Upload Your Datasheet Here...", font=("Helvetica", 24))
label1.grid(row=0, column=0, pady=5, sticky=W, columnspan=3)

select_file_label = Label(frame1, text="Select File : ", font=("Helvetica", 20)).grid(row=2, column=0, sticky=W, padx=10)
select_file_label = Label(frame1, text="Upload to Scrap : ", font=("Helvetica", 20)).grid(row=5, column=0, sticky=W, padx=10)

#frame 2

frame2 = LabelFrame(root, padx=5, pady=5, width=500, height=400)
frame2.grid(row=1, column=0, padx=5, pady=5)
frame2.grid_propagate(False)
label2 = Label(frame2, text="Download Your Datasheet Here...", font=("Helvetica", 24))
label2.grid(row=0, column=0, pady=5, sticky=W, columnspan=3)


label2 = Label(frame2, text="Enter name of CSV data file (branch_sem_year)")
label2.grid(row=3, column=0, sticky=W, padx = 10)
               

entry= Entry(frame2, text='branch_sem_year', width = 20, font=("Helvetica", 16))
entry.grid(row = 4, column = 0)
entry.get()


quit_btn = PhotoImage(file='<address of cancel button image>')
img_lable5 = Label(image=quit_btn)
quit_Button = Button(frame2, image = quit_btn, borderwidth=0, command=my_delete)
quit_Button.grid(row=6, column=0, pady=50)



download_btn = PhotoImage(file='<address of download button image>')
img_lable4 = Label(image=download_btn)
Download_Button2 = Button(frame2, image=download_btn, borderwidth=0, command=lambda:write_to_csv(result))
Download_Button2.grid(row=4, column=1, pady=10)
select_file_label = Label(frame2, text="Download DATA : ", font=("Helvetica", 20)).grid(row=2, column=0, sticky=W, padx=10, pady = 20)


root.mainloop()
